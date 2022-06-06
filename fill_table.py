import sys
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from os.path import exists
import json

def find_element_if_exists(driver, by, name):
    try:
        return driver.find_element(by, name)
    except NoSuchElementException:
        return None

def transform(obj, f):
    return None if obj is None else f(obj)

if __name__ == '__main__':
    table_name = sys.argv[1]
    authors_dump_name = sys.argv[2]
    currency = sys.argv[3]
    
    wb = openpyxl.load_workbook(filename = table_name) if exists(table_name) else openpyxl.Workbook()
    ws = wb[currency] if currency in wb.sheetnames else wb.create_sheet(currency)

    authors_struct = {'last' : 0, 'authors' : {}}
    if exists(authors_dump_name):
        authors_dump = open(authors_dump_name, 'r')
        authors_struct = json.loads(authors_dump.read())
        authors_dump.close()
    new_id = authors_struct['last']
    authors = authors_struct['authors']
    
    opts = webdriver.ChromeOptions()
    opts.headless = True
    opts.add_experimental_option("excludeSwitches", ["enable-logging"])
    driver = webdriver.Chrome(options = opts)
    
    wd_wait = WebDriverWait(driver, 10)
    
    url = input()
    while url:
        print('.', end = '')

        try:
            driver.get(url)
            wd_wait.until(EC.all_of(
                EC.presence_of_element_located((By.ID, 'article-title')),
                EC.presence_of_element_located((By.CLASS_NAME, 'stat_view_metric')),
                EC.presence_of_element_located((By.CLASS_NAME, 'ioa_text')),
                EC.presence_of_element_located((By.CLASS_NAME, 'date')),
                EC.presence_of_element_located((By.CLASS_NAME, 'time'))
            ))

            title = driver.find_element(By.ID, 'article-title').text
            main_part = driver.find_element(By.CLASS_NAME, 'stat_view_metric')
            paragraphs = main_part.find_elements(By.TAG_NAME, 'p')
            body = '\n'.join(map(lambda p: p.text, paragraphs))
            author = driver.find_element(By.CLASS_NAME, 'ioa_text').text.split('\n')[2]
            date = driver.find_element(By.CLASS_NAME, 'date').text
            time = driver.find_element(By.CLASS_NAME, 'time').text
            relevance = transform(find_element_if_exists(driver, By.CLASS_NAME, 'date_relevance_end hot'), lambda x: x.text)

            author_id = new_id
            if author in authors:
                author_id = authors[author]
            else:
                authors[author] = author_id
                new_id += 1

            ws.append([title, author, author_id, url, relevance, body, date, time])
        
            print('+', end = '')
        
        except:
            print('-', end = '')
        
        finally:
            url = input()
    
    print()
    driver.quit()
    wb.save(table_name)

    authors_struct['last'] = new_id
    authors_dump = open(authors_dump_name, 'w')
    authors_dump.write(json.dumps(authors_struct))
    authors_dump.close()